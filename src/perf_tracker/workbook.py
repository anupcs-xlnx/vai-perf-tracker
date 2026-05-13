from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from perf_tracker.date_labels import is_workbook_date_label

TARGET_SUITE = "VE2_QOR_P0_HW"
HW_SUFFIX_RE = re.compile(r"-hw-[A-Za-z0-9_]+$")
PLACEHOLDERS = {"", "na", "n/a"}
NUMERIC_PLACEHOLDERS = {"-", "tbd"}
HISTORICAL_METRICS = {"aie", "vart", "perftest"}
STATIC_HEADER_MAP = {
    "model": "model_name",
    "section": "section",
    "focus": "focus",
    "gops": "gops",
    "customer": "customer",
    "stamp/tp": "stamp_tp",
    "batch/dp": "batch_dp",
    "vai 6.2 goal": "vai_6_2_goal",
    "npu": "npu",
    "6.1 npu": "npu_6_1",
    "vart latency": "vart_latency",
}


@dataclass(frozen=True)
class ModelRecord:
    model_name: str
    section: str | None
    focus: str | None
    gops: float | None
    customer: str | None
    stamp_tp: str | None
    batch_dp: str | None
    vai_6_2_goal: float | None
    npu: float | None
    npu_6_1: float | None
    vart_latency: float | None


@dataclass(frozen=True)
class MeasurementRecord:
    model_name: str
    date_label: str
    aie: float | None
    vart: float | None
    perftest: float | None


@dataclass(frozen=True)
class ParsedWorkbook:
    models: list[ModelRecord]
    measurements: list[MeasurementRecord]
    model_to_test_name: dict[str, str]


def parse_workbook(path: str | Path, *, sheet_name: str | None = None) -> ParsedWorkbook:
    workbook = load_workbook(filename=Path(path), data_only=True)
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet_name}")
        return _parse_dashboard_baselines_sheet(workbook[sheet_name])
    if "Dashboard_Baselines" in workbook.sheetnames:
        return _parse_dashboard_baselines_sheet(workbook["Dashboard_Baselines"])
    latency_sheet = workbook["Latency"]
    models, measurements = _parse_latency_sheet(latency_sheet)
    model_to_test_name = _parse_config_mapping(workbook)
    return ParsedWorkbook(
        models=models,
        measurements=measurements,
        model_to_test_name=model_to_test_name,
    )


def _parse_dashboard_baselines_sheet(sheet: Worksheet) -> ParsedWorkbook:
    header_lookup = {
        _normalize_header(sheet.cell(row=1, column=column_index).value): column_index
        for column_index in range(1, sheet.max_column + 1)
    }
    required_columns = {
        "model": header_lookup.get("model"),
        "customer": header_lookup.get("customer"),
        "target latency": header_lookup.get("target latency"),
        "vai 6.1 latency": header_lookup.get("vai 6.1 latency"),
    }
    missing = [name for name, column in required_columns.items() if column is None]
    if missing:
        raise ValueError(
            f"Dashboard_Baselines sheet is missing required columns: {', '.join(missing)}"
        )
    focus_column = header_lookup.get("focus") or header_lookup.get("task")

    models: list[ModelRecord] = []
    model_to_test_name: dict[str, str] = {}
    for row_index in range(2, sheet.max_row + 1):
        test_name = _normalize_text(sheet.cell(row=row_index, column=required_columns["model"]).value)
        model_name = normalize_test_name(test_name)
        if test_name is None or model_name is None:
            continue
        model_to_test_name[model_name] = test_name
        models.append(
            ModelRecord(
                model_name=model_name,
                section=_normalize_text(
                    sheet.cell(row=row_index, column=required_columns["customer"]).value
                ),
                focus=_normalize_text(sheet.cell(row=row_index, column=focus_column).value)
                if focus_column is not None
                else None,
                gops=None,
                customer=_normalize_text(
                    sheet.cell(row=row_index, column=required_columns["customer"]).value
                ),
                stamp_tp=None,
                batch_dp=None,
                vai_6_2_goal=_normalize_float(
                    sheet.cell(row=row_index, column=required_columns["target latency"]).value
                ),
                npu=None,
                npu_6_1=_normalize_float(
                    sheet.cell(row=row_index, column=required_columns["vai 6.1 latency"]).value
                ),
                vart_latency=None,
            )
        )
    return ParsedWorkbook(models=models, measurements=[], model_to_test_name=model_to_test_name)


def normalize_test_name(test_name: str | None) -> str | None:
    normalized = _normalize_text(test_name)
    if normalized is None:
        return None
    return HW_SUFFIX_RE.sub("", normalized)


def _parse_latency_sheet(sheet: Worksheet) -> tuple[list[ModelRecord], list[MeasurementRecord]]:
    static_columns, historical_groups = _parse_latency_headers(sheet)
    models: list[ModelRecord] = []
    measurements: list[MeasurementRecord] = []

    for row_index in range(3, sheet.max_row + 1):
        model_name = _normalize_text(sheet.cell(row=row_index, column=static_columns["model_name"]).value)
        if model_name is None:
            continue

        models.append(
            ModelRecord(
                model_name=model_name,
                section=_normalize_text(_cell_value(sheet, row_index, static_columns, "section")),
                focus=_normalize_text(_cell_value(sheet, row_index, static_columns, "focus")),
                gops=_normalize_float(_cell_value(sheet, row_index, static_columns, "gops")),
                customer=_normalize_text(_cell_value(sheet, row_index, static_columns, "customer")),
                stamp_tp=_normalize_text(_cell_value(sheet, row_index, static_columns, "stamp_tp")),
                batch_dp=_normalize_text(_cell_value(sheet, row_index, static_columns, "batch_dp")),
                vai_6_2_goal=_normalize_float(
                    _cell_value(sheet, row_index, static_columns, "vai_6_2_goal")
                ),
                npu=_normalize_float(_cell_value(sheet, row_index, static_columns, "npu")),
                npu_6_1=_normalize_float(_cell_value(sheet, row_index, static_columns, "npu_6_1")),
                vart_latency=_normalize_float(
                    _cell_value(sheet, row_index, static_columns, "vart_latency")
                ),
            )
        )

        for date_label, columns in historical_groups.items():
            aie = _normalize_float(_cell_value(sheet, row_index, columns, "aie"))
            vart = _normalize_float(_cell_value(sheet, row_index, columns, "vart"))
            perftest = _normalize_float(_cell_value(sheet, row_index, columns, "perftest"))
            if aie is None and vart is None and perftest is None:
                continue
            measurements.append(
                MeasurementRecord(
                    model_name=model_name,
                    date_label=date_label,
                    aie=aie,
                    vart=vart,
                    perftest=perftest,
                )
            )

    return models, measurements


def _parse_latency_headers(sheet: Worksheet) -> tuple[dict[str, int], dict[str, dict[str, int]]]:
    static_columns: dict[str, int] = {}
    historical_groups: dict[str, dict[str, int]] = {}
    current_date_label: str | None = None

    for column_index in range(1, sheet.max_column + 1):
        top_header = _normalize_text(sheet.cell(row=1, column=column_index).value)
        bottom_header = _normalize_text(sheet.cell(row=2, column=column_index).value)

        if top_header is not None:
            current_date_label = top_header if is_workbook_date_label(top_header) else None

        metric_label = _normalize_header(bottom_header)
        if current_date_label and metric_label in HISTORICAL_METRICS:
            historical_groups.setdefault(current_date_label, {})[metric_label] = column_index
            continue

        for header_candidate in _static_header_candidates(top_header, bottom_header):
            mapped_field = STATIC_HEADER_MAP.get(header_candidate)
            if mapped_field is not None:
                static_columns[mapped_field] = column_index
                break

    missing = [field for field in ("model_name",) if field not in static_columns]
    if missing:
        raise ValueError(f"Latency sheet is missing required columns: {', '.join(missing)}")

    return static_columns, historical_groups


def _parse_config_mapping(workbook: Workbook) -> dict[str, str]:
    if "config-4-29" in workbook.sheetnames:
        return _parse_config_sheet(workbook["config-4-29"], require_target_suite=True)
    if "config" in workbook.sheetnames:
        return _parse_config_sheet(workbook["config"], require_target_suite=False)
    return {}


def _parse_config_sheet(sheet: Worksheet, *, require_target_suite: bool) -> dict[str, str]:
    header_lookup: dict[str, int] = {}
    for column_index in range(1, sheet.max_column + 1):
        header = _normalize_header(sheet.cell(row=1, column=column_index).value)
        if header:
            header_lookup[header] = column_index

    test_name_column = header_lookup.get("test name")
    suite_column = header_lookup.get("suite")
    if test_name_column is None:
        return {}

    model_to_test_name: dict[str, str] = {}
    for row_index in range(2, sheet.max_row + 1):
        test_name = _normalize_text(sheet.cell(row=row_index, column=test_name_column).value)
        if test_name is None:
            continue

        suite_name = None
        if suite_column is not None:
            suite_name = _normalize_text(sheet.cell(row=row_index, column=suite_column).value)

        if require_target_suite and suite_name != TARGET_SUITE:
            continue
        model_name = normalize_test_name(test_name)
        if model_name is not None:
            model_to_test_name[model_name] = test_name

    return model_to_test_name


def _cell_value(sheet: Worksheet, row_index: int, columns: dict[str, int], key: str) -> Any:
    column_index = columns.get(key)
    if column_index is None:
        return None
    return sheet.cell(row=row_index, column=column_index).value


def _combine_headers(top_header: str | None, bottom_header: str | None) -> str:
    parts = [part for part in (top_header, bottom_header) if part]
    return _normalize_header(" ".join(parts))


def _static_header_candidates(top_header: str | None, bottom_header: str | None) -> tuple[str, ...]:
    return (
        _combine_headers(top_header, bottom_header),
        _normalize_header(bottom_header),
        _normalize_header(top_header),
    )


def _normalize_header(value: Any) -> str:
    text = _normalize_text(value)
    if text is None:
        return ""
    return re.sub(r"\s+", " ", text).strip().lower()


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        if normalized.lower() in PLACEHOLDERS:
            return None
        return normalized
    return str(value).strip()


def _normalize_float(value: Any) -> float | None:
    normalized = _normalize_text(value)
    if normalized is None:
        return None
    cleaned = normalized.replace(",", "")
    if cleaned.lower() in NUMERIC_PLACEHOLDERS:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None
