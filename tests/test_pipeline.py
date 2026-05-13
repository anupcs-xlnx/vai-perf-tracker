from __future__ import annotations

from datetime import date
import importlib
import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

REAL_XOAH_SUMMARY_URL = (
    "http://xoah/summary?canned=true&key=VAI_6_2_GEN2_REGRESSION&"
    "param=ACAS&category=VAI-ML&newhome=true&"
    "superSuiteName=VAI_6_2_GEN2_REGRESSION&user=z1aiebuild&relBranch=RAI_1.8"
)


def _load_module(module_name: str):
    try:
        return importlib.import_module(module_name)
    except ModuleNotFoundError as exc:
        pytest.fail(f"Expected {module_name} to exist, but import failed: {exc}")


def _load_config_module():
    return _load_module("perf_tracker.config")


def _load_pipeline_module():
    return _load_module("perf_tracker.pipeline")


def _load_history_module():
    return _load_module("perf_tracker.history")


def _load_xoah_module():
    return _load_module("perf_tracker.xoah")


def _write_workbook(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Latency"

    static_headers = {
        1: "Model",
        2: "Section",
        3: "Focus",
        4: "GOPS ",
        5: "Customer",
        7: "Stamp/tp",
        8: "Batch/dp",
        9: "VAI 6.2 Goal",
        10: "NPU",
        16: "6.1 NPU",
        18: "VART",
    }
    for column, label in static_headers.items():
        sheet.cell(row=1, column=column, value=label)
    sheet.cell(row=2, column=18, value="Latency")

    sheet.merge_cells(start_row=1, start_column=24, end_row=1, end_column=26)
    sheet.cell(row=1, column=24, value="7th May")
    sheet.cell(row=2, column=24, value="AIE")
    sheet.cell(row=2, column=25, value="VART")
    sheet.cell(row=2, column=26, value="Perftest")

    sheet.merge_cells(start_row=1, start_column=27, end_row=1, end_column=29)
    sheet.cell(row=1, column=27, value="5th May")
    sheet.cell(row=2, column=27, value="AIE")
    sheet.cell(row=2, column=28, value="VART")
    sheet.cell(row=2, column=29, value="Perftest")

    row_values = {
        1: "asura_int8_subaru_2x8_t2d1",
        2: "CV",
        3: "Latency",
        4: 123.4,
        5: "Subaru",
        7: "t2d1",
        8: "2x8",
        9: 1.7,
        10: 1.5,
        16: 1.9,
        24: 1.1,
        25: 1.3,
        26: 1.4,
        27: 1.5,
        28: 1.6,
        29: 1.7,
    }
    for column, value in row_values.items():
        sheet.cell(row=3, column=column, value=value)

    config_sheet = workbook.create_sheet("config-4-29")
    config_sheet.append(["Suite", "Test Name"])
    config_sheet.append(["VE2_QOR_P0_HW", "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml"])

    workbook.save(path)
    return path


def _write_dashboard_baselines_workbook(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    p0_sheet = workbook.active
    p0_sheet.title = "Dashboard_QOR_P0_Baselines"
    p0_sheet.append(
        [
            "Model",
            "Display Model",
            "Customer",
            "Focus",
            "Target Latency",
            "VAI 6.1 Latency",
        ]
    )
    p0_sheet.append(
        [
            "p0_model-hw-vek385_vaiml",
            "p0_model",
            "P0 Customer",
            "P0",
            10,
            11,
        ]
    )

    o3_sheet = workbook.create_sheet("Dashboard_O3_Baselines")
    o3_sheet.append(
        [
            "Model",
            "Display Model",
            "Customer",
            "Focus",
            "Target Latency",
            "VAI 6.1 Latency",
        ]
    )
    o3_sheet.append(
        [
            "o3_model-hw-vek385_vaiml",
            "o3_model",
            "O3 Customer",
            "O3",
            20,
            21,
        ]
    )

    workbook.save(path)
    return path


def _write_config(
    path: Path,
    *,
    workbook_path: str,
    history_csv_path: str,
    dashboard_output_dir: str,
    workbook_history_year: int = 2026,
    xoah_summary_url: str = REAL_XOAH_SUMMARY_URL,
    suite_name: str = "VE2_QOR_P0_HW",
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "workbook_path": workbook_path,
                "workbook_history_year": workbook_history_year,
                "xoah_summary_url": xoah_summary_url,
                "suite_name": suite_name,
                "history_csv_path": history_csv_path,
                "dashboard_output_dir": dashboard_output_dir,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return path


def _write_multi_suite_config(
    path: Path,
    *,
    workbook_path: str,
    legacy_history_csv_path: str,
    dashboard_output_dir: str,
    p0_history_csv_path: str,
    o3_history_csv_path: str,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "workbook_path": workbook_path,
                "workbook_history_year": 2026,
                "xoah_summary_url": REAL_XOAH_SUMMARY_URL,
                "history_csv_path": legacy_history_csv_path,
                "dashboard_output_dir": dashboard_output_dir,
                "suites": [
                    {
                        "name": "VE2_QOR_P0_HW",
                        "workbook_sheet": "Dashboard_QOR_P0_Baselines",
                        "history_csv_path": p0_history_csv_path,
                    },
                    {
                        "name": "VE2_QOR_O3_HW",
                        "workbook_sheet": "Dashboard_O3_Baselines",
                        "history_csv_path": o3_history_csv_path,
                    },
                ],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return path


def test_load_tracking_config_resolves_paths_relative_to_config_file(tmp_path: Path) -> None:
    config_module = _load_config_module()
    workbook_path = tmp_path / "inputs" / "latency.xlsx"
    workbook_path.parent.mkdir(parents=True)
    workbook_path.touch()
    config_path = _write_config(
        tmp_path / "config" / "tracking.json",
        workbook_path="../inputs/latency.xlsx",
        history_csv_path="../artifacts/history/history.csv",
        dashboard_output_dir="../artifacts/dashboard",
    )

    config = config_module.load_tracking_config(config_path)

    assert config.config_path == config_path.resolve()
    assert config.workbook_path == workbook_path.resolve()
    assert config.workbook_history_year == 2026
    assert config.xoah_summary_url == REAL_XOAH_SUMMARY_URL
    assert config.suite_name == "VE2_QOR_P0_HW"
    assert config.history_csv_path == (tmp_path / "artifacts" / "history" / "history.csv").resolve()
    assert config.dashboard_output_dir == (tmp_path / "artifacts" / "dashboard").resolve()


def test_load_tracking_config_supports_multi_suite_dashboard_config(tmp_path: Path) -> None:
    config_module = _load_config_module()
    workbook_path = tmp_path / "inputs" / "baselines.xlsx"
    workbook_path.parent.mkdir(parents=True)
    workbook_path.touch()
    config_path = _write_multi_suite_config(
        tmp_path / "config" / "tracking.json",
        workbook_path="../inputs/baselines.xlsx",
        legacy_history_csv_path="../artifacts/history/history.csv",
        dashboard_output_dir="../artifacts/dashboard",
        p0_history_csv_path="../artifacts/history/VE2_QOR_P0_HW.csv",
        o3_history_csv_path="../artifacts/history/VE2_QOR_O3_HW.csv",
    )

    config = config_module.load_tracking_config(config_path)

    assert [suite.name for suite in config.suites] == ["VE2_QOR_P0_HW", "VE2_QOR_O3_HW"]
    assert [suite.workbook_sheet for suite in config.suites] == [
        "Dashboard_QOR_P0_Baselines",
        "Dashboard_O3_Baselines",
    ]
    assert config.suites[0].history_csv_path == (
        tmp_path / "artifacts" / "history" / "VE2_QOR_P0_HW.csv"
    ).resolve()
    assert config.suites[1].history_csv_path == (
        tmp_path / "artifacts" / "history" / "VE2_QOR_O3_HW.csv"
    ).resolve()


@pytest.mark.parametrize(
    ("field_name", "field_value"),
    [
        ("workbook_path", None),
        ("history_csv_path", None),
        ("dashboard_output_dir", None),
        ("xoah_summary_url", None),
        ("suite_name", None),
    ],
)
def test_load_tracking_config_rejects_non_string_text_and_path_fields(
    tmp_path: Path,
    field_name: str,
    field_value: object,
) -> None:
    config_module = _load_config_module()
    config_path = tmp_path / "config" / "tracking.json"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(
        json.dumps(
            {
                "workbook_path": "../inputs/latency.xlsx",
                "workbook_history_year": 2026,
                "xoah_summary_url": REAL_XOAH_SUMMARY_URL,
                "suite_name": "VE2_QOR_P0_HW",
                "history_csv_path": "../artifacts/history/history.csv",
                "dashboard_output_dir": "../artifacts/dashboard",
                field_name: field_value,
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match=field_name):
        config_module.load_tracking_config(config_path)


def test_run_pipeline_writes_canonical_history_and_dashboard_outputs(tmp_path: Path) -> None:
    config_module = _load_config_module()
    pipeline_module = _load_pipeline_module()
    history = _load_history_module()
    xoah = _load_xoah_module()

    workbook_path = _write_workbook(tmp_path / "inputs" / "latency.xlsx")
    config = config_module.load_tracking_config(
        _write_config(
            tmp_path / "config" / "tracking.json",
            workbook_path="../inputs/latency.xlsx",
            history_csv_path="../artifacts/history/history.csv",
            dashboard_output_dir="../artifacts/dashboard",
        )
    )

    extractor_calls: list[tuple[str, str, str]] = []

    def fake_nightly_extractor(parsed_workbook, source, *, suite_name: str, cached_suite_run_names):
        extractor_calls.append((source.super_suite, source.user, suite_name))
        assert cached_suite_run_names == set()
        assert parsed_workbook.model_to_test_name == {
            "asura_int8_subaru_2x8_t2d1": "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml"
        }
        return [
            xoah.NightlyRecord(
                model_name="asura_int8_subaru_2x8_t2d1",
                test_name="asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
                suite_name=suite_name,
                suite_run_name="20260507_004500_VE2_QOR_P0_HW",
                user=source.user,
                rel_branch=source.rel_branch,
                super_suite=source.super_suite,
                vart_latency_ms=0.9,
            )
        ]

    result = pipeline_module.run_pipeline(config, nightly_extractor=fake_nightly_extractor)

    assert workbook_path.exists()
    assert extractor_calls == [("VAI_6_2_GEN2_REGRESSION", "z1aiebuild", "VE2_QOR_P0_HW")]
    assert result.history_csv_path == config.history_csv_path
    assert result.latest_dashboard_path == config.dashboard_output_dir / "latest.html"
    assert result.workbook_seeded_row_count == 0
    assert result.prior_xoah_row_count == 0
    assert result.current_xoah_row_count == 1
    assert result.canonical_row_count == 1
    assert result.snapshot_paths == [
        config.dashboard_output_dir / "daily" / "2026-05-07.html",
    ]
    assert all(path.exists() for path in result.snapshot_paths)
    assert result.latest_dashboard_path.exists()

    written_rows = history.read_history_csv(config.history_csv_path)
    by_date = {row.date: row for row in written_rows}
    assert set(by_date) == {date(2026, 5, 7)}

    may_7 = by_date[date(2026, 5, 7)]
    assert may_7.source_kind == "xoah"
    assert may_7.vart_latency_ms == pytest.approx(0.9)
    assert may_7.aie_latency_ms is None
    assert may_7.perftest_latency_ms is None


def test_run_pipeline_writes_flat_dashboards_for_each_configured_suite_and_migrates_p0_history(
    tmp_path: Path,
) -> None:
    config_module = _load_config_module()
    pipeline_module = _load_pipeline_module()
    history = _load_history_module()
    xoah = _load_xoah_module()

    config = config_module.load_tracking_config(
        _write_multi_suite_config(
            tmp_path / "config" / "tracking.json",
            workbook_path="../inputs/baselines.xlsx",
            legacy_history_csv_path="../artifacts/history/history.csv",
            dashboard_output_dir="../artifacts/dashboard",
            p0_history_csv_path="../artifacts/history/VE2_QOR_P0_HW.csv",
            o3_history_csv_path="../artifacts/history/VE2_QOR_O3_HW.csv",
        )
    )
    _write_dashboard_baselines_workbook(config.workbook_path)
    config.history_csv_path.parent.mkdir(parents=True, exist_ok=True)
    history.write_history_csv(
        config.history_csv_path,
        [
            history.HistoryRow(
                date=date(2026, 5, 6),
                model_name="p0_model",
                section="Old P0 Customer",
                focus="Old P0",
                customer="Old P0 Customer",
                gops=None,
                stamp_tp=None,
                batch_dp=None,
                target_latency_ms=9.0,
                npu_latency_ms=None,
                vai61_latency_ms=8.0,
                vart_latency_ms=7.0,
                aie_latency_ms=None,
                perftest_latency_ms=None,
                source_kind="xoah",
                suite_run_name="20260506_004500_VE2_QOR_P0_HW",
                suite_name="VE2_QOR_P0_HW",
                user="z1aiebuild",
                rel_branch="RAI_1.8",
                super_suite="VAI_6_2_GEN2_REGRESSION",
                test_name="p0_model-hw-vek385_vaiml",
            )
        ],
    )

    extractor_calls: list[tuple[str, set[str], tuple[str, ...]]] = []

    def fake_nightly_extractor(parsed_workbook, source, *, suite_name: str, cached_suite_run_names):
        extractor_calls.append(
            (
                suite_name,
                set(cached_suite_run_names),
                tuple(model.model_name for model in parsed_workbook.models),
            )
        )
        if suite_name == "VE2_QOR_P0_HW":
            return [
                xoah.NightlyRecord(
                    model_name="p0_model",
                    test_name="p0_model-hw-vek385_vaiml",
                    suite_name=suite_name,
                    suite_run_name="20260507_004500_VE2_QOR_P0_HW",
                    user=source.user,
                    rel_branch=source.rel_branch,
                    super_suite=source.super_suite,
                    vart_latency_ms=6.0,
                )
            ]
        return [
            xoah.NightlyRecord(
                model_name="o3_model",
                test_name="o3_model-hw-vek385_vaiml",
                suite_name=suite_name,
                suite_run_name="20260507_004500_VE2_QOR_O3_HW",
                user=source.user,
                rel_branch=source.rel_branch,
                super_suite=source.super_suite,
                vart_latency_ms=16.0,
            )
        ]

    result = pipeline_module.run_pipeline(config, nightly_extractor=fake_nightly_extractor)

    assert extractor_calls == [
        (
            "VE2_QOR_P0_HW",
            {"20260506_004500_VE2_QOR_P0_HW"},
            ("p0_model",),
        ),
        ("VE2_QOR_O3_HW", set(), ("o3_model",)),
    ]
    suite_results = {suite_result.suite_name: suite_result for suite_result in result.suite_results}
    assert set(suite_results) == {"VE2_QOR_P0_HW", "VE2_QOR_O3_HW"}

    p0_rows = history.read_history_csv(tmp_path / "artifacts" / "history" / "VE2_QOR_P0_HW.csv")
    o3_rows = history.read_history_csv(tmp_path / "artifacts" / "history" / "VE2_QOR_O3_HW.csv")
    assert [(row.date, row.vart_latency_ms, row.customer) for row in p0_rows] == [
        (date(2026, 5, 6), 7.0, "P0 Customer"),
        (date(2026, 5, 7), 6.0, "P0 Customer"),
    ]
    assert [(row.date, row.vart_latency_ms, row.customer) for row in o3_rows] == [
        (date(2026, 5, 7), 16.0, "O3 Customer")
    ]

    dashboard_dir = tmp_path / "artifacts" / "dashboard"
    assert suite_results["VE2_QOR_P0_HW"].snapshot_paths == [
        dashboard_dir / "VE2_QOR_P0_HW" / "2026-05-06.html",
        dashboard_dir / "VE2_QOR_P0_HW" / "2026-05-07.html",
    ]
    assert suite_results["VE2_QOR_O3_HW"].snapshot_paths == [
        dashboard_dir / "VE2_QOR_O3_HW" / "2026-05-07.html"
    ]
    assert (dashboard_dir / "VE2_QOR_P0_HW_latest.html").is_symlink()
    assert (dashboard_dir / "VE2_QOR_P0_HW_latest.html").readlink() == Path(
        "VE2_QOR_P0_HW/2026-05-07.html"
    )
    assert (dashboard_dir / "VE2_QOR_O3_HW_latest.html").is_symlink()
    assert (dashboard_dir / "VE2_QOR_O3_HW_latest.html").readlink() == Path(
        "VE2_QOR_O3_HW/2026-05-07.html"
    )
    assert not (dashboard_dir / "latest.html").exists()
    assert not list(dashboard_dir.glob("VE2_QOR_*_20*.html"))
    assert not (dashboard_dir / "daily").exists()


def test_run_pipeline_caches_all_xoah_rows_but_displays_workbook_models_only(tmp_path: Path) -> None:
    config_module = _load_config_module()
    pipeline_module = _load_pipeline_module()
    history = _load_history_module()
    xoah = _load_xoah_module()

    config = config_module.load_tracking_config(
        _write_config(
            tmp_path / "config" / "tracking.json",
            workbook_path="../inputs/latency.xlsx",
            history_csv_path="../artifacts/history/history.csv",
            dashboard_output_dir="../artifacts/dashboard",
        )
    )
    _write_workbook(config.workbook_path)
    workbook = load_workbook(config.workbook_path)
    latency_sheet = workbook["Latency"]
    latency_sheet.cell(row=4, column=1, value="display_only_model")
    latency_sheet.cell(row=4, column=2, value="Display Customer")
    latency_sheet.cell(row=4, column=9, value=9.0)
    latency_sheet.cell(row=4, column=16, value=10.0)
    workbook.save(config.workbook_path)

    def fake_nightly_extractor(parsed_workbook, source, *, suite_name: str, cached_suite_run_names):
        return [
            xoah.NightlyRecord(
                model_name="asura_int8_subaru_2x8_t2d1",
                test_name="asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
                suite_name=suite_name,
                suite_run_name="20260507_004500_VE2_QOR_P0_HW",
                user=source.user,
                rel_branch=source.rel_branch,
                super_suite=source.super_suite,
                vart_latency_ms=0.9,
            ),
            xoah.NightlyRecord(
                model_name="not_in_display_workbook",
                test_name="not_in_display_workbook-hw-vek385_vaiml",
                suite_name=suite_name,
                suite_run_name="20260507_004500_VE2_QOR_P0_HW",
                user=source.user,
                rel_branch=source.rel_branch,
                super_suite=source.super_suite,
                vart_latency_ms=3.3,
            ),
        ]

    result = pipeline_module.run_pipeline(config, nightly_extractor=fake_nightly_extractor)

    cached_rows = history.read_history_csv(config.history_csv_path)
    assert {row.model_name for row in cached_rows} == {
        "asura_int8_subaru_2x8_t2d1",
        "not_in_display_workbook",
    }
    latest_html = result.latest_dashboard_path.read_text(encoding="utf-8")
    assert "asura_int8_subaru_2x8_t2d1" in latest_html
    assert "display_only_model" in latest_html
    assert "No cached XOAH data for this model" in latest_html
    assert "not_in_display_workbook" not in latest_html


def test_run_pipeline_preserves_existing_xoah_rows_while_reseeding_workbook_rows(
    tmp_path: Path,
) -> None:
    config_module = _load_config_module()
    pipeline_module = _load_pipeline_module()
    history = _load_history_module()
    xoah = _load_xoah_module()

    config = config_module.load_tracking_config(
        _write_config(
            tmp_path / "config" / "tracking.json",
            workbook_path="../inputs/latency.xlsx",
            history_csv_path="../artifacts/history/history.csv",
            dashboard_output_dir="../artifacts/dashboard",
        )
    )
    _write_workbook(config.workbook_path)
    config.history_csv_path.parent.mkdir(parents=True, exist_ok=True)

    history.write_history_csv(
        config.history_csv_path,
        [
            history.HistoryRow(
                date=date(2026, 5, 5),
                model_name="asura_int8_subaru_2x8_t2d1",
                section="CV",
                focus="Latency",
                customer="Subaru",
                gops=123.4,
                stamp_tp="t2d1",
                batch_dp="2x8",
                target_latency_ms=1.7,
                npu_latency_ms=1.5,
                vai61_latency_ms=1.9,
                vart_latency_ms=99.0,
                aie_latency_ms=99.0,
                perftest_latency_ms=99.0,
                source_kind="workbook",
                suite_run_name=None,
                suite_name=None,
                user=None,
                rel_branch=None,
                super_suite=None,
                test_name=None,
            ),
            history.HistoryRow(
                date=date(2026, 5, 6),
                model_name="asura_int8_subaru_2x8_t2d1",
                section="CV",
                focus="Latency",
                customer="Subaru",
                gops=123.4,
                stamp_tp="t2d1",
                batch_dp="2x8",
                target_latency_ms=1.7,
                npu_latency_ms=1.5,
                vai61_latency_ms=1.9,
                vart_latency_ms=1.05,
                aie_latency_ms=None,
                perftest_latency_ms=None,
                source_kind="xoah",
                suite_run_name="20260506_004500_VE2_QOR_P0_HW",
                suite_name="VE2_QOR_P0_HW",
                user="z1aiebuild",
                rel_branch="RAI_1.8",
                super_suite="VAI_6_2_GEN2_REGRESSION",
                test_name="asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
            ),
        ],
    )

    def fake_nightly_extractor(parsed_workbook, source, *, suite_name: str, cached_suite_run_names):
        assert cached_suite_run_names == {"20260506_004500_VE2_QOR_P0_HW"}
        return [
            xoah.NightlyRecord(
                model_name="asura_int8_subaru_2x8_t2d1",
                test_name="asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
                suite_name=suite_name,
                suite_run_name="20260507_004500_VE2_QOR_P0_HW",
                user=source.user,
                rel_branch=source.rel_branch,
                super_suite=source.super_suite,
                vart_latency_ms=0.9,
            )
        ]

    result = pipeline_module.run_pipeline(config, nightly_extractor=fake_nightly_extractor)
    written_rows = history.read_history_csv(config.history_csv_path)
    by_date = {row.date: row for row in written_rows}

    assert result.prior_xoah_row_count == 1
    assert set(by_date) == {date(2026, 5, 6), date(2026, 5, 7)}
    assert by_date[date(2026, 5, 6)].source_kind == "xoah"
    assert by_date[date(2026, 5, 6)].vart_latency_ms == pytest.approx(1.05)
    assert by_date[date(2026, 5, 7)].source_kind == "xoah"
    assert by_date[date(2026, 5, 7)].vart_latency_ms == pytest.approx(0.9)


def test_run_pipeline_refreshes_preserved_xoah_rows_with_current_workbook_metadata(
    tmp_path: Path,
) -> None:
    config_module = _load_config_module()
    pipeline_module = _load_pipeline_module()
    history = _load_history_module()

    config = config_module.load_tracking_config(
        _write_config(
            tmp_path / "config" / "tracking.json",
            workbook_path="../inputs/latency.xlsx",
            history_csv_path="../artifacts/history/history.csv",
            dashboard_output_dir="../artifacts/dashboard",
        )
    )
    _write_workbook(config.workbook_path)
    config.history_csv_path.parent.mkdir(parents=True, exist_ok=True)

    history.write_history_csv(
        config.history_csv_path,
        [
            history.HistoryRow(
                date=date(2026, 5, 6),
                model_name="asura_int8_subaru_2x8_t2d1",
                section="Old Section",
                focus="Old Focus",
                customer="Old Customer",
                gops=10.0,
                stamp_tp="legacy-stamp",
                batch_dp="legacy-batch",
                target_latency_ms=9.9,
                npu_latency_ms=8.8,
                vai61_latency_ms=7.7,
                vart_latency_ms=1.05,
                aie_latency_ms=None,
                perftest_latency_ms=None,
                source_kind="xoah",
                suite_run_name="20260506_004500_VE2_QOR_P0_HW",
                suite_name="VE2_QOR_P0_HW",
                user="z1aiebuild",
                rel_branch="RAI_1.8",
                super_suite="VAI_6_2_GEN2_REGRESSION",
                test_name="asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
            )
        ],
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Latency"
    static_headers = {
        1: "Model",
        2: "Section",
        3: "Focus",
        4: "GOPS ",
        5: "Customer",
        7: "Stamp/tp",
        8: "Batch/dp",
        9: "VAI 6.2 Goal",
        10: "NPU",
        16: "6.1 NPU",
        18: "VART",
    }
    for column, label in static_headers.items():
        sheet.cell(row=1, column=column, value=label)
    sheet.cell(row=2, column=18, value="Latency")
    sheet.merge_cells(start_row=1, start_column=24, end_row=1, end_column=26)
    sheet.cell(row=1, column=24, value="7th May")
    sheet.cell(row=2, column=24, value="AIE")
    sheet.cell(row=2, column=25, value="VART")
    sheet.cell(row=2, column=26, value="Perftest")
    updated_values = {
        1: "asura_int8_subaru_2x8_t2d1",
        2: "Updated Section",
        3: "Updated Focus",
        4: 222.0,
        5: "Updated Customer",
        7: "t4d2",
        8: "4x2",
        9: 1.1,
        10: 0.8,
        16: 1.2,
        24: 1.5,
        25: 1.6,
        26: 1.7,
    }
    for column, value in updated_values.items():
        sheet.cell(row=3, column=column, value=value)
    config_sheet = workbook.create_sheet("config-4-29")
    config_sheet.append(["Suite", "Test Name"])
    config_sheet.append(["VE2_QOR_P0_HW", "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml"])
    workbook.save(config.workbook_path)

    result = pipeline_module.run_pipeline(config, nightly_extractor=lambda *_args, **_kwargs: [])
    written_rows = history.read_history_csv(config.history_csv_path)
    by_date = {row.date: row for row in written_rows}

    assert result.prior_xoah_row_count == 1
    preserved_row = by_date[date(2026, 5, 6)]
    assert preserved_row.source_kind == "xoah"
    assert preserved_row.section == "Updated Section"
    assert preserved_row.focus == "Updated Focus"
    assert preserved_row.customer == "Updated Customer"
    assert preserved_row.gops == pytest.approx(222.0)
    assert preserved_row.stamp_tp == "t4d2"
    assert preserved_row.batch_dp == "4x2"
    assert preserved_row.target_latency_ms == pytest.approx(1.1)
    assert preserved_row.npu_latency_ms == pytest.approx(0.8)
    assert preserved_row.vai61_latency_ms == pytest.approx(1.2)
    assert preserved_row.vart_latency_ms == pytest.approx(1.05)
    assert preserved_row.suite_run_name == "20260506_004500_VE2_QOR_P0_HW"
    assert preserved_row.user == "z1aiebuild"
