from __future__ import annotations

import importlib
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))


def _load_workbook_module():
    try:
        return importlib.import_module("perf_tracker.workbook")
    except ModuleNotFoundError as exc:
        pytest.fail(f"Expected perf_tracker.workbook to exist, but import failed: {exc}")


def _parse_workbook(path: Path):
    module = _load_workbook_module()
    parse_workbook = getattr(module, "parse_workbook", None)
    if parse_workbook is None:
        pytest.fail("Expected perf_tracker.workbook.parse_workbook to be defined.")
    return parse_workbook(path)


def _parse_workbook_sheet(path: Path, sheet_name: str):
    module = _load_workbook_module()
    parse_workbook = getattr(module, "parse_workbook", None)
    if parse_workbook is None:
        pytest.fail("Expected perf_tracker.workbook.parse_workbook to be defined.")
    return parse_workbook(path, sheet_name=sheet_name)


def _build_latency_sheet(
    workbook: Workbook,
    *,
    use_spreadsheet_numeric_variants: bool = False,
    use_realistic_combined_headers: bool = False,
    history_date_labels: tuple[str, str] = ("7th May", "5th May"),
) -> None:
    sheet = workbook.active
    sheet.title = "Latency"

    static_headers = {
        1: "Model",
        2: "Section",
        3: "Focus",
        5: "Customer",
        7: "Stamp/tp",
        8: "Batch/dp",
    }
    for column, label in static_headers.items():
        sheet.cell(row=1, column=column, value=label)

    if use_realistic_combined_headers:
        sheet.cell(row=1, column=4, value="GOPS")
        sheet.cell(row=2, column=4, value="GOPS ")
        sheet.cell(row=1, column=9, value="Latency")
        sheet.cell(row=2, column=9, value="VAI 6.2 Goal")
        sheet.cell(row=1, column=10, value="VAI 6.2 Latency")
        sheet.cell(row=2, column=10, value="NPU")
        sheet.cell(row=1, column=16, value="6.1_XOAH")
        sheet.cell(row=2, column=16, value="6.1 NPU")
        sheet.cell(row=1, column=18, value="VART")
        sheet.cell(row=2, column=18, value="Latency")
    else:
        sheet.cell(row=1, column=4, value="GOPS ")
        sheet.cell(row=1, column=9, value="VAI 6.2 Goal")
        sheet.cell(row=1, column=10, value="NPU")
        sheet.cell(row=1, column=16, value="6.1 NPU")
        sheet.cell(row=1, column=18, value="VART")
        sheet.cell(row=2, column=18, value="Latency")

    sheet.merge_cells(start_row=1, start_column=24, end_row=1, end_column=26)
    sheet.cell(row=1, column=24, value=history_date_labels[0])
    sheet.cell(row=2, column=24, value="AIE")
    sheet.cell(row=2, column=25, value="VART")
    sheet.cell(row=2, column=26, value="Perftest")

    sheet.merge_cells(start_row=1, start_column=27, end_row=1, end_column=29)
    sheet.cell(row=1, column=27, value=history_date_labels[1])
    sheet.cell(row=2, column=27, value="AIE")
    sheet.cell(row=2, column=28, value="VART")
    sheet.cell(row=2, column=29, value="Perftest")

    first_row = {
        1: "asura_int8_subaru_2x8_t2d1",
        2: "CV",
        3: "Latency",
        4: 123.4,
        5: "Subaru",
        7: "t2d1",
        8: "2x8",
        9: 1.7,
        10: 1.5,
        16: "NA",
        18: "N/A",
        24: 1.1,
        25: 1.3,
        26: "",
        27: "NA",
        28: 1.4,
        29: 1.6,
    }
    second_row = {
        1: "phoenix_int8_city",
        2: "NLP",
        3: "N/A",
        4: "",
        5: "",
        7: "batch1",
        8: "1x1",
        9: 2.2,
        10: 2.0,
        16: 2.6,
        18: 2.8,
        24: 2.1,
        25: 2.3,
        26: 2.4,
        27: "",
        28: "",
        29: "",
    }
    if use_spreadsheet_numeric_variants:
        first_row[4] = "1,234.5"
        first_row[16] = "-"
        first_row[18] = "TBD"
        first_row[24] = "-"
        first_row[25] = "1,234.5"
        first_row[26] = "TBD"

    for column, value in first_row.items():
        sheet.cell(row=3, column=column, value=value)
    for column, value in second_row.items():
        sheet.cell(row=4, column=column, value=value)


def _build_config_4_29_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("config-4-29")
    sheet.append(["Suite", "Test Name"])
    sheet.append(["VE2_QOR_P0_HW", "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml"])
    sheet.append(["OTHER_SUITE", "ignored_model-hw-vek385_vaiml"])
    sheet.append(["VE2_QOR_P0_HW", "phoenix_int8_city-hw-vek385_vaiml"])


def _build_config_fallback_sheet(workbook: Workbook, *, include_suite_column: bool = False) -> None:
    sheet = workbook.create_sheet("config")
    if include_suite_column:
        sheet.append(["Suite", "Test Name"])
        sheet.append(["OTHER_SUITE", "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml"])
        sheet.append(["ANOTHER_SUITE", "fallback_model-hw-vek385_vaiml"])
        return

    sheet.append(["Test Name"])
    sheet.append(["asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml"])
    sheet.append(["fallback_model-hw-vek385_vaiml"])


def _write_workbook(
    path: Path,
    *,
    include_config_4_29: bool,
    fallback_has_suite_column: bool = False,
    use_spreadsheet_numeric_variants: bool = False,
    use_realistic_combined_headers: bool = False,
    history_date_labels: tuple[str, str] = ("7th May", "5th May"),
) -> Path:
    workbook = Workbook()
    _build_latency_sheet(
        workbook,
        use_spreadsheet_numeric_variants=use_spreadsheet_numeric_variants,
        use_realistic_combined_headers=use_realistic_combined_headers,
        history_date_labels=history_date_labels,
    )
    if include_config_4_29:
        _build_config_4_29_sheet(workbook)
    else:
        _build_config_fallback_sheet(
            workbook,
            include_suite_column=fallback_has_suite_column,
        )
    workbook.save(path)
    return path


def test_parse_workbook_extracts_models_measurements_and_suite_mapping(tmp_path: Path) -> None:
    workbook_path = _write_workbook(tmp_path / "latency.xlsx", include_config_4_29=True)

    parsed = _parse_workbook(workbook_path)

    assert [record.model_name for record in parsed.models] == [
        "asura_int8_subaru_2x8_t2d1",
        "phoenix_int8_city",
    ]

    first_model = parsed.models[0]
    assert first_model.section == "CV"
    assert first_model.focus == "Latency"
    assert first_model.gops == 123.4
    assert first_model.customer == "Subaru"
    assert first_model.stamp_tp == "t2d1"
    assert first_model.batch_dp == "2x8"
    assert first_model.vai_6_2_goal == 1.7
    assert first_model.npu == 1.5
    assert first_model.npu_6_1 is None
    assert first_model.vart_latency is None

    second_model = parsed.models[1]
    assert second_model.focus is None
    assert second_model.gops is None
    assert second_model.customer is None
    assert second_model.npu_6_1 == 2.6
    assert second_model.vart_latency == 2.8

    measurement_tuples = {
        (
            measurement.model_name,
            measurement.date_label,
            measurement.aie,
            measurement.vart,
            measurement.perftest,
        )
        for measurement in parsed.measurements
    }
    assert measurement_tuples == {
        ("asura_int8_subaru_2x8_t2d1", "7th May", 1.1, 1.3, None),
        ("asura_int8_subaru_2x8_t2d1", "5th May", None, 1.4, 1.6),
        ("phoenix_int8_city", "7th May", 2.1, 2.3, 2.4),
    }

    assert parsed.model_to_test_name == {
        "asura_int8_subaru_2x8_t2d1": "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
        "phoenix_int8_city": "phoenix_int8_city-hw-vek385_vaiml",
    }


def test_parse_workbook_falls_back_to_config_sheet_when_config_4_29_is_missing(
    tmp_path: Path,
) -> None:
    workbook_path = _write_workbook(tmp_path / "latency.xlsx", include_config_4_29=False)

    parsed = _parse_workbook(workbook_path)

    assert parsed.model_to_test_name == {
        "asura_int8_subaru_2x8_t2d1": "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
        "fallback_model": "fallback_model-hw-vek385_vaiml",
    }


def test_parse_workbook_fallback_config_ignores_suite_values_when_config_4_29_is_missing(
    tmp_path: Path,
) -> None:
    workbook_path = _write_workbook(
        tmp_path / "latency.xlsx",
        include_config_4_29=False,
        fallback_has_suite_column=True,
    )

    parsed = _parse_workbook(workbook_path)

    assert parsed.model_to_test_name == {
        "asura_int8_subaru_2x8_t2d1": "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
        "fallback_model": "fallback_model-hw-vek385_vaiml",
    }


def test_parse_workbook_normalizes_common_spreadsheet_numeric_values(tmp_path: Path) -> None:
    workbook_path = _write_workbook(
        tmp_path / "latency.xlsx",
        include_config_4_29=True,
        use_spreadsheet_numeric_variants=True,
    )

    try:
        parsed = _parse_workbook(workbook_path)
    except ValueError as exc:
        pytest.fail(
            "Expected spreadsheet numeric placeholders to normalize without crashing, "
            f"but parsing raised: {exc}"
        )

    first_model = parsed.models[0]
    assert first_model.gops == 1234.5
    assert first_model.npu_6_1 is None
    assert first_model.vart_latency is None

    first_model_measurements = {
        measurement.date_label: measurement
        for measurement in parsed.measurements
        if measurement.model_name == "asura_int8_subaru_2x8_t2d1"
    }
    assert first_model_measurements["7th May"].aie is None
    assert first_model_measurements["7th May"].vart == 1234.5
    assert first_model_measurements["7th May"].perftest is None


def test_parse_workbook_supports_dashboard_baselines_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard_Baselines"
    sheet.append(
        [
            "Model",
            "Display Model",
            "Customer",
            "Focus",
            "Target Latency",
            "VAI 6.1 Latency",
        ]
    )
    sheet.append(
        [
            "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml",
            "asura_int8_subaru_2x8_t2d1",
            "Subaru",
            "P0",
            18,
            "24.47",
        ]
    )
    workbook_path = tmp_path / "baselines.xlsx"
    workbook.save(workbook_path)

    parsed = _parse_workbook(workbook_path)

    assert parsed.measurements == []
    assert parsed.model_to_test_name == {
        "asura_int8_subaru_2x8_t2d1": "asura_int8_subaru_2x8_t2d1-hw-vek385_vaiml"
    }
    assert parsed.models[0].model_name == "asura_int8_subaru_2x8_t2d1"
    assert parsed.models[0].section == "Subaru"
    assert parsed.models[0].customer == "Subaru"
    assert parsed.models[0].focus == "P0"
    assert parsed.models[0].vai_6_2_goal == 18.0
    assert parsed.models[0].npu_6_1 == 24.47


def test_parse_workbook_selects_named_dashboard_baseline_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    p0_sheet = workbook.active
    p0_sheet.title = "Dashboard_QOR_P0_Baselines"
    p0_sheet.append(
        [
            "Model",
            "Display Model",
            "Customer",
            "Focus",
            "Target Latency",
            "VAI 6.1 Latency",
        ]
    )
    p0_sheet.append(
        [
            "p0_model-hw-vek385_vaiml",
            "p0_model",
            "P0 Customer",
            "P0",
            10,
            11,
        ]
    )
    o3_sheet = workbook.create_sheet("Dashboard_O3_Baselines")
    o3_sheet.append(
        [
            "Model",
            "Display Model",
            "Customer",
            "Focus",
            "Target Latency",
            "VAI 6.1 Latency",
        ]
    )
    o3_sheet.append(
        [
            "o3_model-hw-vek385_vaiml",
            "o3_model",
            "O3 Customer",
            "O3",
            20,
            21,
        ]
    )
    workbook_path = tmp_path / "baselines.xlsx"
    workbook.save(workbook_path)

    parsed = _parse_workbook_sheet(workbook_path, "Dashboard_O3_Baselines")

    assert [model.model_name for model in parsed.models] == ["o3_model"]
    assert parsed.model_to_test_name == {"o3_model": "o3_model-hw-vek385_vaiml"}
    assert parsed.models[0].section == "O3 Customer"
    assert parsed.models[0].focus == "O3"
    assert parsed.models[0].vai_6_2_goal == 20.0
    assert parsed.models[0].npu_6_1 == 21.0


def test_parse_workbook_uses_task_as_focus_when_dashboard_sheet_has_no_focus(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard_O3_Baselines"
    sheet.append(
        [
            "Model",
            "Task",
            "Customer",
            "Orin FPS",
            "Target Latency",
            "VAI 6.1 Latency",
            "Notes",
        ]
    )
    sheet.append(
        [
            "o3_model-hw-vek385_vaiml",
            "O3 task group",
            "O3 Customer",
            30,
            20,
            21,
            "tracked",
        ]
    )
    workbook_path = tmp_path / "baselines.xlsx"
    workbook.save(workbook_path)

    parsed = _parse_workbook_sheet(workbook_path, "Dashboard_O3_Baselines")

    assert parsed.models[0].model_name == "o3_model"
    assert parsed.models[0].section == "O3 Customer"
    assert parsed.models[0].customer == "O3 Customer"
    assert parsed.models[0].focus == "O3 task group"


def test_parse_workbook_handles_realistic_combined_latency_headers(tmp_path: Path) -> None:
    workbook_path = _write_workbook(
        tmp_path / "latency.xlsx",
        include_config_4_29=True,
        use_realistic_combined_headers=True,
    )

    parsed = _parse_workbook(workbook_path)

    first_model = parsed.models[0]
    assert first_model.gops == 123.4
    assert first_model.vai_6_2_goal == 1.7
    assert first_model.npu == 1.5
    assert first_model.npu_6_1 is None
    assert first_model.vart_latency is None

    second_model = parsed.models[1]
    assert second_model.gops is None
    assert second_model.vai_6_2_goal == 2.2
    assert second_model.npu == 2.0
    assert second_model.npu_6_1 == 2.6
    assert second_model.vart_latency == 2.8


def test_parse_workbook_ignores_invalid_history_month_labels(tmp_path: Path) -> None:
    workbook_path = _write_workbook(
        tmp_path / "latency.xlsx",
        include_config_4_29=True,
        history_date_labels=("7th May", "5th Smarch"),
    )

    parsed = _parse_workbook(workbook_path)

    measurement_tuples = {
        (
            measurement.model_name,
            measurement.date_label,
            measurement.aie,
            measurement.vart,
            measurement.perftest,
        )
        for measurement in parsed.measurements
    }
    assert measurement_tuples == {
        ("asura_int8_subaru_2x8_t2d1", "7th May", 1.1, 1.3, None),
        ("phoenix_int8_city", "7th May", 2.1, 2.3, 2.4),
    }
